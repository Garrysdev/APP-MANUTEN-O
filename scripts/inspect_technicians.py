import openpyxl
import pyxlsb

def inspect_files():
    f1 = r'C:\Users\Quinta do Arrobe\Desktop\DOWNLOADS CHROME\drive-download-20260725T123638Z-1-001\PL-MAN-01 PLANO MANUTENÇÃO_2026.xlsx'
    wb1 = openpyxl.load_workbook(f1, data_only=True)
    print("=== SHEETS EM PL-MAN-01 ===")
    print(wb1.sheetnames)
    for sheet_name in wb1.sheetnames:
        s = wb1[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (primeiras 5 linhas) ---")
        for r in range(1, 6):
            row_vals = [cell for cell in next(s.iter_rows(min_row=r, max_row=r, values_only=True)) if cell is not None]
            if row_vals:
                print(f"L{r}: {row_vals[:10]}")

    f2 = r'C:\Users\Quinta do Arrobe\Desktop\DOWNLOADS CHROME\drive-download-20260725T123638Z-1-001\FR-MAN-09 MANUTENÇÃO_05_01_2026_8.xlsb'
    wb2 = pyxlsb.open_workbook(f2)
    print("\n=== SHEETS EM FR-MAN-09 ===")
    print(wb2.sheets)
    for sheet_name in wb2.sheets:
        print(f"\n--- Sheet: {sheet_name} (primeiras 5 linhas) ---")
        with wb2.get_sheet(sheet_name) as s:
            rows = list(s.rows())
            for r in range(min(5, len(rows))):
                row_vals = [cell.v for cell in rows[r] if cell.v is not None]
                if row_vals:
                    print(f"L{r+1}: {row_vals[:10]}")

if __name__ == '__main__':
    inspect_files()
